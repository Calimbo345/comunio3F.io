import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

MAX_JUGADORES = 14


def crear_liga(nombre):
    archivo_excel = f"{nombre}.xlsx"

    if os.path.exists(archivo_excel):
        print(f"La liga '{nombre}' ya existe.")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Liga"

    # Configurar la hoja
    ws['A1'] = "Jugadores"
    ws['A1'].font = Font(bold=True)
    ws['E1'] = f"Calendario_liga_{nombre}"
    ws['E1'].font = Font(bold=True)

    # Preparar celdas para jugadores
    for i in range(2, MAX_JUGADORES + 2):
        ws.cell(row=i, column=1)

    # Guardar el archivo
    wb.save(archivo_excel)
    print(f"Liga '{nombre}' creada con éxito.")
    return True


def unirse_a_liga(usuario, nombre_liga):
    archivo_excel = f"{nombre_liga}.xlsx"

    if not os.path.exists(archivo_excel):
        print(f"La liga '{nombre_liga}' no existe.")
        return False

    wb = load_workbook(archivo_excel)
    ws = wb.active

    jugadores_actuales = []
    for row in ws.iter_rows(min_row=2, max_row=MAX_JUGADORES + 1, min_col=1, max_col=1):
        for cell in row:
            if cell.value:
                jugadores_actuales.append(cell.value)

    if len(jugadores_actuales) >= MAX_JUGADORES:
        print(f"La liga '{nombre_liga}' ya tiene el máximo de {MAX_JUGADORES} jugadores.")
        return False

    if usuario in jugadores_actuales:
        print(f"{usuario} ya está en la liga '{nombre_liga}'.")
        return False

    # Añadir el nombre del jugador en un hueco vacío
    for row in range(2, MAX_JUGADORES + 2):
        if ws.cell(row=row, column=1).value is None:
            ws.cell(row=row, column=1, value=usuario)
            break

    # Sustituir "Descanso" por el nombre del nuevo usuario en el calendario
    for row in range(2, ws.max_row + 1):
        for col in range(5, 9):  # Columnas E a H
            if ws.cell(row=row, column=col).value == "Descanso":
                ws.cell(row=row, column=col, value=usuario)
                break  # Solo sustituir la primera ocurrencia de "Descanso"

    wb.save(archivo_excel)
    print(f"{usuario} se ha unido a la liga '{nombre_liga}'.")

    # Regenerar el calendario si es necesario
    if len(jugadores_actuales) + 1 >= 2:
        generar_calendario(nombre_liga)

    return True


def generar_calendario(nombre_liga):
    archivo_excel = f"{nombre_liga}.xlsx"
    wb = load_workbook(archivo_excel)
    ws = wb.active

    jugadores = [cell.value for cell in ws['A'] if cell.value and cell.row > 1]
    equipos = jugadores.copy()
    while len(equipos) < MAX_JUGADORES:
        equipos.append("Descanso")

    n = len(equipos)
    calendario = []
    jornada = []
    for i in range(n - 1):
        for j in range(n // 2):
            local = equipos[j]
            visitante = equipos[n - 1 - j]
            jornada.append((local, visitante))
        calendario.append(jornada)
        equipos = [equipos[0]] + [equipos[-1]] + equipos[1:-1]

    # Generar jornadas de vuelta
    jornadas_vuelta = [(visitante, local) for jornada in calendario for local, visitante in jornada]
    calendario.extend([jornadas_vuelta[i:i + len(jornada)] for i in range(0, len(jornadas_vuelta), len(jornada))])

    # Escribir el calendario en el Excel
    row = 2
    for i, jornada in enumerate(calendario, 1):
        ws.cell(row=row, column=4, value=f"Jornada {i}")  # Jornada en columna D
        ws.cell(row=row, column=4).font = Font(bold=True)
        row += 1
        for partido in jornada:
            ws.cell(row=row, column=5, value=partido[0])  # Local en columna E
            ws.cell(row=row, column=6, value="0")  # Puntos local en columna F
            ws.cell(row=row, column=7, value="0")  # Puntos visitante en columna G
            ws.cell(row=row, column=8, value=partido[1])  # Visitante en columna H
            row += 1
        row += 1  # Añadir una línea en blanco entre jornadas

    # Ajustar el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(archivo_excel)
    print(f"Calendario generado y guardado para la liga '{nombre_liga}'.")


def actualizar_resultado(nombre_liga, usuario1, usuario2, pts1, pts2, jornada_param):
    global local, visitante
    archivo_excel = f"{nombre_liga}.xlsx"

    if not os.path.exists(archivo_excel):
        print(f"La liga '{nombre_liga}' no existe.")
        return False, None, None

    try:
        wb = load_workbook(archivo_excel)
        ws = wb.active

        exito = False
        ganador = None
        perdedor = None
        nombre_jornada = f"Jornada {jornada_param}"

        print(f"Buscando partido entre {usuario1} y {usuario2} en la {nombre_jornada}")

        jornada_encontrada = False
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=4).value == nombre_jornada:  # Cambiado a columna D
                jornada_encontrada = True
                continue

            if jornada_encontrada:
                local = ws.cell(row=row, column=5).value
                visitante = ws.cell(row=row, column=8).value

                if local is None or visitante is None:
                    break

                if (local == usuario1 and visitante == usuario2) or (local == usuario2 and visitante == usuario1):
                    exito = True

                    # Si uno de los equipos es "Descanso", establecer resultado a "0-0"
                    if local == "Descanso" or visitante == "Descanso":
                        pts1 = pts2 = 0

                    # Actualizar los puntos
                    if local == usuario1:
                        ws.cell(row=row, column=6, value=pts1)
                        ws.cell(row=row, column=7, value=pts2)
                    else:
                        ws.cell(row=row, column=6, value=pts2)
                        ws.cell(row=row, column=7, value=pts1)

                    ganador, perdedor = (usuario1, usuario2) if pts1 > pts2 else (
                        usuario2, usuario1) if pts2 > pts1 else ("Empate", "Empate")

                    print(f"Partido encontrado: {local} vs {visitante}, Resultado: {pts1}-{pts2}")
                    break

        if exito:
            wb.save(archivo_excel)
            print(f"Archivo guardado: {archivo_excel}")
            print(f"Resultado actualizado: {local} vs {visitante}, Resultado: {pts1}-{pts2}")
        else:
            print(f"No se encontró un partido entre {usuario1} y {usuario2} para la {nombre_jornada}.")

        return exito, ganador, perdedor
    except Exception as e:
        print(f"Error al actualizar los resultados: {e}")
        return False, None, None


def mostrar_enfrentamientos(nombre_liga, jornada):
    archivo_excel = f"{nombre_liga}.xlsx"
    if not os.path.exists(archivo_excel):
        print(f"La liga '{nombre_liga}' no existe.")
        return None

    wb = load_workbook(archivo_excel)
    ws = wb.active

    jornada_encontrada = False
    enfrentamientos = []

    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=4).value  # Cambiado a columna D
        if cell_value == f"Jornada {jornada}":
            jornada_encontrada = True
            continue

        if jornada_encontrada:
            local = ws.cell(row=row, column=5).value
            visitante = ws.cell(row=row, column=8).value
            pts_local = ws.cell(row=row, column=6).value
            pts_visitante = ws.cell(row=row, column=7).value

            if local is None or visitante is None:  # Fin de la jornada
                break

            enfrentamientos.append((local, visitante, pts_local, pts_visitante))

    if not jornada_encontrada:
        print(f"No se encontró la jornada {jornada} en la liga '{nombre_liga}'.")
        return None
    elif not enfrentamientos:
        print(f"No hay enfrentamientos programados para la jornada {jornada} en la liga '{nombre_liga}'.")
        return []
    else:
        return enfrentamientos



